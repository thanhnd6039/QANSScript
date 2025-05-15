from openpyxl import load_workbook
# from pyexcel.cookbook import merge_all_to_a_book
import glob
import os
from robot.api import logger
import pandas as pd

class CExcel(object):
    def create_excel_file(self, filePath):
        df = pd.DataFrame()
        df.to_excel(filePath, index=False)

    def get_number_of_cols_in_excel(self, filePath):
        numOfCols = 0
        file = load_workbook(filePath)
        sheet = file.active
        numOfCols = sheet.max_column
        return numOfCols
    def get_number_of_rows_in_excel(self, filePath, sheetName=''):
        numOfRows = 0
        file = load_workbook(filePath)
        if sheetName == '':
            sheet = file.active
        else:
            sheet = file[sheetName]
        numOfRows = sheet.max_row
        return numOfRows
    def convert_csv_to_xlsx(self, csvFilePath, xlsxFilePath):
        # logger.console(csvFilePath)
        # logger.console(xlsxFilePath)
        df = pd.read_csv(csvFilePath)
        df.to_excel(xlsxFilePath, index=False)
        os.remove(csvFilePath)

    def get_position_of_column(self, filePath, rowIndex, searchStr):
        posOfColumn = 0
        numOfCols = self.get_number_of_cols_in_excel(filePath)
        file = load_workbook(filePath)
        sheet = file.active
        # valueOfCell = sheet.cell(row=3, column=6).value
        # logger.console("valueOfCell: {0}".format(valueOfCell))
        rowIndex = int(rowIndex)
        for colIndex in range(1, numOfCols+1):
            valueOfCell = sheet.cell(row=rowIndex, column=colIndex).value
            if searchStr == valueOfCell:
                posOfColumn = colIndex
                break
        return posOfColumn



# if __name__ == '__main__':
#     cExcel = CExcel()
#     filePath = 'C:\\RobotFramework\\Downloads\\Sales Gap Report NS With SO Forecast.xlsx'
#     rowIndex = 3
#     searchStr = '2024.Q1 R'
#     number = cExcel.get_position_of_column(filePath, rowIndex, searchStr)
#     print(number)



